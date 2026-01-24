"""
Export Service

Module 5: Dashboard, Analytics & Monitoring
Service for exporting data to CSV and Excel formats.
"""

import csv
import io
from typing import List, Dict, Any
from datetime import datetime
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from src.models.campaign import Campaign
from src.models.call_session import CallSession
from src.models.lead import Lead
from src.utils.logger import StructuredLogger

logger = StructuredLogger(__name__)


class ExportService:
    """
    Service for exporting data to CSV/Excel
    """

    def __init__(self, db_session: AsyncSession):
        self.db = db_session

    async def export_campaign_results(
        self,
        campaign_id: int,
        format: str = "csv"
    ) -> bytes:
        """
        Export campaign results with all call outcomes
        """
        # Get campaign
        result = await self.db.execute(
            select(Campaign).where(Campaign.id == campaign_id)
        )
        campaign = result.scalar_one_or_none()

        if not campaign:
            raise ValueError("Campaign not found")

        # Get all leads with their call results
        leads_result = await self.db.execute(
            select(Lead)
            .where(Lead.campaign_id == campaign_id)
            .order_by(Lead.id)
        )
        leads = leads_result.scalars().all()

        # Build data
        data = []
        for lead in leads:
            # Get last call
            call_result = await self.db.execute(
                select(CallSession)
                .where(CallSession.lead_id == lead.id)
                .order_by(CallSession.initiated_at.desc())
                .limit(1)
            )
            last_call = call_result.scalar_one_or_none()

            row = {
                "Lead Name": lead.name,
                "Phone": lead.phone,
                "Email": lead.email or "",
                "Property Type": lead.property_type,
                "Location": lead.location,
                "Budget": lead.budget or "",
                "Source": lead.source,
                "Call Attempts": lead.call_attempts,
                "Last Called": lead.last_call_attempt.strftime("%Y-%m-%d %H:%M") if lead.last_call_attempt else "",
                "Outcome": last_call.outcome.value if last_call and last_call.outcome else "Not Called",
                "Duration (sec)": last_call.duration_seconds if last_call else "",
                "Recording URL": last_call.recording_url if last_call else ""
            }

            # Add collected data if available
            if last_call and last_call.collected_data:
                try:
                    collected = json.loads(last_call.collected_data)
                    row["Purpose"] = collected.get("purpose", "")
                    row["Timeline"] = collected.get("timeline", "")
                    row["Interested"] = collected.get("interested", "")
                except (json.JSONDecodeError, TypeError):
                    row["Purpose"] = ""
                    row["Timeline"] = ""
                    row["Interested"] = ""
            else:
                row["Purpose"] = ""
                row["Timeline"] = ""
                row["Interested"] = ""

            data.append(row)

        if format == "csv":
            return self._export_to_csv(data)
        else:
            return self._export_to_excel(data, f"Campaign_{campaign.name}_Results")

    async def export_call_transcripts(
        self,
        campaign_id: int,
        format: str = "csv"
    ) -> bytes:
        """
        Export call transcripts
        """
        # Get all call sessions for campaign
        result = await self.db.execute(
            select(CallSession)
            .join(Lead)
            .where(Lead.campaign_id == campaign_id)
            .order_by(CallSession.initiated_at.desc())
        )
        call_sessions = result.scalars().all()

        data = []
        for call in call_sessions:
            transcript_text = ""
            if call.full_transcript:
                try:
                    transcript = json.loads(call.full_transcript)
                    transcript_text = "\n".join([
                        f"{t.get('speaker', 'unknown').upper()}: {t.get('text', '')}"
                        for t in transcript
                    ])
                except (json.JSONDecodeError, TypeError):
                    transcript_text = "Error parsing transcript"

            data.append({
                "Call SID": call.call_sid,
                "Lead Name": call.lead.name,
                "Phone": call.lead.phone,
                "Date": call.initiated_at.strftime("%Y-%m-%d %H:%M"),
                "Duration (sec)": call.duration_seconds or 0,
                "Outcome": call.outcome.value if call.outcome else "",
                "Transcript": transcript_text
            })

        if format == "csv":
            return self._export_to_csv(data)
        else:
            return self._export_to_excel(data, "Call_Transcripts")

    async def export_lead_list(
        self,
        campaign_id: int,
        format: str = "csv"
    ) -> bytes:
        """
        Export simple lead list
        """
        # Get all leads
        result = await self.db.execute(
            select(Lead)
            .where(Lead.campaign_id == campaign_id)
            .order_by(Lead.id)
        )
        leads = result.scalars().all()

        data = []
        for lead in leads:
            data.append({
                "ID": lead.id,
                "Name": lead.name,
                "Phone": lead.phone,
                "Email": lead.email or "",
                "Property Type": lead.property_type,
                "Location": lead.location,
                "Budget": lead.budget or "",
                "Source": lead.source,
                "Call Attempts": lead.call_attempts
            })

        if format == "csv":
            return self._export_to_csv(data)
        else:
            return self._export_to_excel(data, "Lead_List")

    def _export_to_csv(self, data: List[Dict[str, Any]]) -> bytes:
        """Export data to CSV format"""
        if not data:
            return b""

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

        return output.getvalue().encode('utf-8')

    def _export_to_excel(self, data: List[Dict[str, Any]], sheet_name: str) -> bytes:
        """Export data to Excel format"""
        if not data:
            # Return empty workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name limit is 31 characters

        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                value = row_data[header]
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            # Set width with a max limit
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()
